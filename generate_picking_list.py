import pandas as pd
import os
import glob
import unicodedata
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import re

# ==========================================
# 設定エリア
# ==========================================
INPUT_DIR = "input"
OUTPUT_DIR = "output"

# 列の位置設定
IDX_JAN = 16
IDX_STORE_START = 37

# 各項目の列番号（0始まり）
COL_IDX_DELIVERY_DATE = 3  # D列
COL_IDX_CLIENT_NAME = 8  # I列
COL_IDX_MK_CODE = 14  # O列
COL_IDX_PROD_NAME = 15  # P列
COL_IDX_COLOR_CODE = 21  # W列
COL_IDX_COLOR_NAME = 23  # X列
COL_IDX_SIZE_NAME = 25  # Z列
COL_IDX_ORDER_QTY = 26  # AA列
COL_IDX_JAN = 16  # Q列

# 行の設定（0始まり）
ROW_IDX_STORE_INFO = 0  # 1行目
ROW_IDX_DATA_START = 1  # 2行目

# 処理を停止するキーワード
STOP_COLUMN_KEYWORD = "伝票枝番"

# 【サイズ順序定義】
SIZE_ORDER = [
    "XS",
    "SS",
    "S",
    "M",
    "L",
    "LL",
    "XL",
    "3L",
    "4L",
    "5L",
    "FREE",
    "F",
    "Free",
    "OneSize",
]


def sanitize_sheet_name(name):
    """Excelのシート名に使えない文字を除去・置換し、31文字以内に収める"""
    if pd.isna(name) or name == "":
        return "Unknown"
    name = re.sub(r"[\\|/|:|?|*|[|]]", "_", str(name))
    return str(name).strip()[:31]


def get_str_width(s):
    """文字列の表示幅計算（列幅自動調整用）"""
    width = 0
    for c in str(s):
        if unicodedata.east_asian_width(c) in ("F", "W", "A"):
            width += 2
        else:
            width += 1
    return width


def format_worksheet(writer, sheet_name, title_text, subtitle_text=None, header_row=3):
    """
    Excelの見た目を整える総合関数
    - タイトル・サブタイトルの挿入
    - 罫線・列幅調整
    - 印刷設定
    """
    ws = writer.sheets[sheet_name]

    # 1. タイトルとサブタイトルの書き込み
    ws["A1"] = title_text
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left")

    if subtitle_text:
        ws["A2"] = subtitle_text
        ws["A2"].font = Font(bold=True, size=11)
        ws["A2"].alignment = Alignment(horizontal="left")

    # 2. スタイル定義
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ヘッダー行以降（データ部分）のスタイル適用
    max_column_widths = {}

    for row in ws.iter_rows(min_row=header_row):
        for cell in row:
            # 基本スタイル
            cell.alignment = Alignment(
                vertical="top", horizontal="left", wrap_text=True
            )
            cell.border = thin_border

            # 列幅計算
            if cell.value:
                w = get_str_width(cell.value)
                col_idx = cell.column
                if w > max_column_widths.get(col_idx, 0):
                    max_column_widths[col_idx] = w

    # 3. ヘッダー行（header_row）のスタイル
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )

    # 4. 列幅の適用
    for col_idx, width in max_column_widths.items():
        column_letter = get_column_letter(col_idx)
        adjusted_width = min((width * 1.2) + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 5. 「チェック」列の特別対応
    check_col_idx = None
    for cell in ws[header_row]:
        if cell.value == "チェック":
            check_col_idx = cell.column
            break

    if check_col_idx:
        check_letter = get_column_letter(check_col_idx)
        ws.column_dimensions[check_letter].width = 8
        for row in ws.iter_rows(
            min_row=header_row + 1, min_col=check_col_idx, max_col=check_col_idx
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 6. 印刷設定
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False

    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3


def normalize_size_text(text):
    """サイズ文字正規化"""
    if pd.isna(text):
        return ""
    s = str(text)
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"[\s　]*(サイズ|size|Size)[\s　]*", "", s, flags=re.IGNORECASE)
    return s.strip()


def sort_df_with_custom_size(df, sort_cols, size_col_name="サイズ"):
    """サイズ順ソート"""
    if df.empty:
        return df
    temp_df = df.copy()
    if size_col_name in sort_cols and size_col_name in temp_df.columns:
        rank_map = {size: i for i, size in enumerate(SIZE_ORDER)}
        temp_df["_normalized_size"] = temp_df[size_col_name].apply(normalize_size_text)
        temp_df["_size_rank"] = temp_df["_normalized_size"].map(rank_map).fillna(999)
        actual_sort_cols = [
            ("_size_rank" if c == size_col_name else c) for c in sort_cols
        ]
        temp_df = temp_df.sort_values(by=actual_sort_cols)
        return temp_df.drop(columns=["_normalized_size", "_size_rank"])
    else:
        return temp_df.sort_values(by=sort_cols)


def process_single_csv(file_path):
    filename = os.path.basename(file_path)
    base_name, _ = os.path.splitext(filename)
    # 出力用サブフォルダを作成
    file_output_dir = os.path.join(OUTPUT_DIR, base_name)
    os.makedirs(file_output_dir, exist_ok=True)
    print(f"処理中: {filename} ...")

    # 1. CSV読み込み
    try:
        try:
            df = pd.read_csv(file_path, encoding="utf-8", header=None)
        except UnicodeDecodeError:
            print("  [情報] Shift-JISで再試行します...")
            df = pd.read_csv(file_path, encoding="cp932", header=None)
    except Exception as e:
        print(f"  [エラー] 読み込み失敗: {e}")
        return

    if len(df.columns) <= 26:
        print("  [スキップ] 列数が不足しています")
        return

    # 2. 基本情報の抽出
    df_data = df.iloc[ROW_IDX_DATA_START:].reset_index(drop=True)
    col_map = {
        "納品日": df_data.iloc[:, COL_IDX_DELIVERY_DATE],
        "得意先（センター）名": df_data.iloc[:, COL_IDX_CLIENT_NAME],
        "MK品番": df_data.iloc[:, COL_IDX_MK_CODE],
        "商品名": df_data.iloc[:, COL_IDX_PROD_NAME],
        "MK_COLOR": df_data.iloc[:, COL_IDX_COLOR_CODE],
        "色名": df_data.iloc[:, COL_IDX_COLOR_NAME],
        "サイズ": df_data.iloc[:, COL_IDX_SIZE_NAME],
        "JAN": df_data.iloc[:, COL_IDX_JAN],
    }

    # 3. 店舗列処理
    store_data_list = []
    total_cols = len(df.columns)
    for i in range(IDX_STORE_START, total_cols, 3):
        if i + 2 >= total_cols:
            break
        current_header = (
            str(df.iloc[ROW_IDX_STORE_INFO, i])
            if not pd.isna(df.iloc[ROW_IDX_STORE_INFO, i])
            else ""
        )
        if STOP_COLUMN_KEYWORD in current_header:
            break

        store_code = (
            str(df.iloc[ROW_IDX_DATA_START, i]).strip()
            if not pd.isna(df.iloc[ROW_IDX_DATA_START, i])
            else ""
        )
        store_name = (
            str(df.iloc[ROW_IDX_DATA_START, i + 1]).strip()
            if not pd.isna(df.iloc[ROW_IDX_DATA_START, i + 1])
            else ""
        )
        qty_series = df.iloc[ROW_IDX_DATA_START:, i + 2]

        block = pd.DataFrame(
            {
                "店舗コード": store_code,
                "店舗名": store_name,
                "数量": qty_series.values,
                "original_index": df_data.index,
            }
        )
        block["数量"] = pd.to_numeric(block["数量"], errors="coerce").fillna(0)
        block = block[block["数量"] > 0]
        if not block.empty:
            store_data_list.append(block)

    if not store_data_list:
        print(f"  [スキップ] 有効なデータがありません: {filename}")
        return

    long_df = pd.concat(store_data_list, ignore_index=True)
    base_df = pd.DataFrame(col_map)
    final_df = pd.concat(
        [
            base_df.iloc[long_df["original_index"]].reset_index(drop=True),
            long_df.reset_index(drop=True),
        ],
        axis=1,
    )

    final_df["サイズ"] = final_df["サイズ"].apply(normalize_size_text)

    # ==========================================
    # 1. 商品毎のピッキングリスト
    # ==========================================
    group_cols_1 = ["商品名", "MK品番", "MK_COLOR", "色名", "サイズ"]
    out1 = final_df.groupby(group_cols_1, as_index=False)["数量"].sum()
    out1 = out1.rename(columns={"MK品番": "品番", "数量": "合計枚数"})

    out1 = sort_df_with_custom_size(out1, ["商品名", "品番", "MK_COLOR", "サイズ"])
    out1["チェック"] = ""
    out1_indexed = out1.set_index(["商品名", "品番", "MK_COLOR", "色名", "サイズ"])

    path1 = os.path.join(file_output_dir, "1_商品別.xlsx")
    with pd.ExcelWriter(path1, engine="openpyxl") as writer:
        out1_indexed.to_excel(
            writer, sheet_name="商品別集計", startrow=2, merge_cells=True
        )
        format_worksheet(
            writer,
            sheet_name="商品別集計",
            title_text="商品別ピッキングリスト",
            header_row=3,
        )
    print(f"  -> 出力1完了: {os.path.basename(path1)}")

    # ==========================================
    # 2. センター毎のピッキングリスト
    # ==========================================
    group_cols_2 = [
        "得意先（センター）名",
        "商品名",
        "MK品番",
        "MK_COLOR",
        "色名",
        "サイズ",
    ]
    out2_base = final_df.groupby(group_cols_2, as_index=False)["数量"].sum()
    out2_base = out2_base.rename(
        columns={
            "MK品番": "品番",
            "数量": "発注数",
            "得意先（センター）名": "得意先（センター名）",
        }
    )

    path2 = os.path.join(file_output_dir, "2_センター別.xlsx")
    with pd.ExcelWriter(path2, engine="openpyxl") as writer:
        centers = out2_base["得意先（センター名）"].unique()
        for center in centers:
            df_center = out2_base[out2_base["得意先（センター名）"] == center].copy()
            df_center = sort_df_with_custom_size(
                df_center, ["商品名", "品番", "MK_COLOR", "サイズ"]
            )

            df_center_out = df_center[
                ["商品名", "品番", "MK_COLOR", "色名", "サイズ", "発注数"]
            ].copy()
            sheet_name = sanitize_sheet_name(center)

            df_center_indexed = df_center_out.set_index(
                ["商品名", "品番", "MK_COLOR", "色名", "サイズ"]
            )
            df_center_indexed.to_excel(
                writer, sheet_name=sheet_name, startrow=2, merge_cells=True
            )

            format_worksheet(
                writer,
                sheet_name=sheet_name,
                title_text="センター別ピッキングリスト",
                subtitle_text=f"得意先（センター）名： {center}",
                header_row=3,
            )
    print(f"  -> 出力2完了: {os.path.basename(path2)}")

    # ==========================================
    # 3. 店舗毎のピッキングリスト
    # ==========================================
    out3_base = final_df.copy()
    out3_base = out3_base.rename(
        columns={"MK品番": "品番", "得意先（センター）名": "得意先名（センター名）"}
    )

    # ソート時に商品名も考慮するように修正
    out3_base = sort_df_with_custom_size(
        out3_base,
        [
            "得意先名（センター名）",
            "店舗コード",
            "店舗名",
            "商品名",
            "品番",
            "MK_COLOR",
            "サイズ",
        ],
    )

    path3 = os.path.join(file_output_dir, "3_店舗別.xlsx")

    try:
        with pd.ExcelWriter(path3, engine="openpyxl") as writer:
            groups = out3_base.groupby(["店舗コード", "店舗名"], sort=False)

            for (st_code, st_name), df_store in groups:
                client_name = df_store["得意先名（センター名）"].iloc[0]

                # ★修正点: 商品名を先頭に追加
                df_out = df_store[["商品名", "品番", "色名", "サイズ", "数量"]].copy()
                total_qty = df_out["数量"].sum()

                # ★修正点: インデックスを設定して書き出すことで、自動的にセル結合させる
                # インデックス: 商品名, 品番, 色名, サイズ
                df_indexed = df_out.set_index(["商品名", "品番", "色名", "サイズ"])

                sheet_name = sanitize_sheet_name(f"{st_code}_{st_name}")

                # データ書き出し (merge_cells=Trueで結合)
                df_indexed.to_excel(
                    writer, sheet_name=sheet_name, startrow=2, merge_cells=True
                )

                # 合計行の追加 (to_excelの後ろに手動で書き込む)
                ws = writer.sheets[sheet_name]
                last_row = ws.max_row + 1

                # 合計ラベル (A列 = 商品名列)
                cell_label = ws.cell(row=last_row, column=1, value="合計")
                cell_label.font = Font(bold=True)

                # 合計値 (E列 = 数量列。インデックス4列 + データ1列 = 5列目)
                cell_val = ws.cell(row=last_row, column=5, value=total_qty)
                cell_val.font = Font(bold=True)

                subtitle = (
                    f"得意先（センター）名：{client_name}   店舗：{st_code} {st_name}"
                )
                format_worksheet(
                    writer,
                    sheet_name=sheet_name,
                    title_text="店舗別ピッキングリスト",
                    subtitle_text=subtitle,
                    header_row=3,
                )

        print(f"  -> 出力3完了: {os.path.basename(path3)}")
    except Exception as e:
        print(f"  [エラー] 出力3作成失敗: {e}")
        import traceback

        traceback.print_exc()


def main():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    files = glob.glob(os.path.join(INPUT_DIR, "*.csv"))
    if not files:
        print(f"'{INPUT_DIR}' フォルダにCSVファイルが見つかりません。")
        return
    print(f"{len(files)} 個のファイルを検出しました。")
    for file_path in files:
        process_single_csv(file_path)
    print("\n処理完了。Enterキーを押して終了してください。")
    input()


if __name__ == "__main__":
    main()
